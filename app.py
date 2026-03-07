import io
import math
import sqlite3
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Лемана Про — юнит-экономика FBS / FBO",
    layout="wide",
    page_icon="📦",
)

DB_PATH = "lemanpro_products.db"


# =========================
# DB
# =========================
def init_db():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    cur = conn.cursor()
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS products (
            sku TEXT PRIMARY KEY,
            name TEXT,
            template TEXT,
            item_type TEXT,
            subcategory TEXT,
            category TEXT,
            length_cm REAL DEFAULT 0,
            width_cm REAL DEFAULT 0,
            height_cm REAL DEFAULT 0,
            weight_kg REAL DEFAULT 0,
            cost_price REAL DEFAULT 0,
            current_price REAL DEFAULT 0,
            promo_price REAL DEFAULT 0,
            region TEXT,
            manual_commission REAL
        )
        """
    )
    conn.commit()
    return conn


# =========================
# Utils
# =========================
def to_float(value, default=0.0):
    if value is None:
        return default
    try:
        if isinstance(value, str):
            value = value.replace("\xa0", "").replace(" ", "").replace(",", ".").strip()
            if value == "":
                return default
        return float(value)
    except Exception:
        return default


def normalize_dimension(raw, unit: str) -> float:
    value = to_float(raw, 0.0)
    unit = (unit or "").strip().lower()
    if unit in ("мм", "mm"):
        return value / 10.0
    if unit in ("м", "meter", "метр", "метры"):
        return value * 100.0
    return value


def normalize_weight(raw, unit: str) -> float:
    value = to_float(raw, 0.0)
    unit = (unit or "").strip().lower()
    if unit in ("г", "гр", "g", "gr"):
        return value / 1000.0
    return value


def first_existing(row: pd.Series, names: List[str], default=None):
    row_map = {str(k).strip().lower(): row[k] for k in row.index}
    for n in names:
        v = row_map.get(n.strip().lower())
        if pd.notna(v):
            return v
    return default


def clean_text(x) -> str:
    if x is None or (isinstance(x, float) and math.isnan(x)):
        return ""
    return str(x).strip()


def liters_from_dimensions(length_cm: float, width_cm: float, height_cm: float) -> float:
    if min(length_cm, width_cm, height_cm) <= 0:
        return 0.0
    return (length_cm * width_cm * height_cm) / 1000.0


def parse_bracket(text: str) -> Tuple[float, Optional[float]]:
    s = clean_text(text).lower().replace(",", ".")
    nums = []
    current = ""
    for ch in s:
        if ch.isdigit() or ch == ".":
            current += ch
        else:
            if current:
                nums.append(float(current))
                current = ""
    if current:
        nums.append(float(current))

    if "от" in s and "+" in s:
        low = nums[0] if nums else 0.0
        return low, None
    if "от" in s and "до" in s and len(nums) >= 2:
        return nums[0], nums[1]
    if len(nums) >= 2:
        return nums[0], nums[1]
    if len(nums) == 1:
        return 0.0, nums[0]
    return 0.0, None


def value_in_bracket(value: float, low: float, high: Optional[float]) -> bool:
    if high is None:
        return value >= low
    return low <= value <= high


def round_to_step(value: float, step: int) -> float:
    if step <= 1:
        return round(value, 2)
    return float(int(math.ceil(max(value, 0.0) / step)) * step)


def safe_mean(values: List[float]) -> float:
    clean = [float(v) for v in values if v is not None]
    return sum(clean) / len(clean) if clean else 0.0


# =========================
# Tariff loading
# =========================
@st.cache_data(show_spinner=False)
def load_commissions(file_path: str) -> Dict[str, Dict[str, float]]:
    xl = pd.ExcelFile(file_path)
    sheet = "Комиссия_FBS и FBO"
    if sheet not in xl.sheet_names:
        raise ValueError(f"В файле нет листа '{sheet}'")

    df = pd.read_excel(file_path, sheet_name=sheet)
    df.columns = [clean_text(c) for c in df.columns]

    result = {
        "template": {},
        "item_type": {},
        "subcategory": {},
        "category": {},
    }

    for _, row in df.iterrows():
        commission_raw = to_float(first_existing(row, ["Комиссия"]), 0.0)
        commission = commission_raw if commission_raw <= 1 else commission_raw / 100.0

        template = clean_text(first_existing(row, ["Шаблон товара"]))
        item_type = clean_text(first_existing(row, ["Тип товара"]))
        subcategory = clean_text(first_existing(row, ["Подкатегория товара"]))
        category = clean_text(first_existing(row, ["Категория"]))

        if template:
            result["template"][template.lower()] = commission
        if item_type:
            result["item_type"][item_type.lower()] = commission
        if subcategory:
            result["subcategory"][subcategory.lower()] = commission
        if category:
            result["category"][category.lower()] = commission

    return result


@st.cache_data(show_spinner=False)
def load_logistics(file_path: str):
    xl = pd.ExcelFile(file_path)

    zero_df = pd.read_excel(file_path, sheet_name="Тарифы (Доставка до СЦ)")
    return_zero_df = pd.read_excel(file_path, sheet_name="Тарифы (Возврат Доставка до СЦ)")
    last_mile_df = pd.read_excel(file_path, sheet_name="Тарифы (Последняя миля)")
    return_last_mile_df = pd.read_excel(file_path, sheet_name="Тарифы (возврат Последняя миля)")
    zones_df = pd.read_excel(file_path, sheet_name="Зоны (услуга Последняя миля)")

    def build_simple_table(df):
        rows = []
        for _, row in df.iterrows():
            bracket = clean_text(first_existing(row, ["Объемный брейк отправления, в л", "Объемный брейк отправления, в л."]))
            tariff = to_float(first_existing(row, ["Тариф, с НДС"]), 0.0)
            low, high = parse_bracket(bracket)
            extra_per_l = tariff if ("от 120" in bracket.lower() and tariff > 0) else 0.0
            base_tariff = 0.0 if ("от 120" in bracket.lower() and tariff > 0) else tariff
            rows.append(
                {
                    "raw_bracket": bracket,
                    "low": low,
                    "high": high,
                    "base_tariff": base_tariff,
                    "extra_per_l": extra_per_l,
                }
            )
        return rows

    def build_last_mile_table(df):
        rows = []
        for _, row in df.iterrows():
            origin_zone = int(to_float(first_existing(row, ["Зона откуда"]), 0))
            dest_label = clean_text(first_existing(row, ["Зона куда"]))
            bracket = clean_text(first_existing(row, ["Весовой брейк отправления, в л.", "Весовой брейк отправления, в л"]))
            tariff = to_float(first_existing(row, ["Тариф, с НДС"]), 0.0)
            extra_per_l = to_float(first_existing(row, ["+1 л, с НДС"]), 0.0)
            low, high = parse_bracket(bracket)
            rows.append(
                {
                    "origin_zone": origin_zone,
                    "dest_label": dest_label,
                    "raw_bracket": bracket,
                    "low": low,
                    "high": high,
                    "base_tariff": tariff,
                    "extra_per_l": extra_per_l,
                }
            )
        return rows

    zone_map = {}
    for _, row in zones_df.iterrows():
        region = clean_text(first_existing(row, ["Край/Область"]))
        zone = int(to_float(first_existing(row, ["Зона"]), 0))
        if region:
            zone_map[region.lower()] = zone

    return {
        "zero_mile": build_simple_table(zero_df),
        "return_zero_mile": build_simple_table(return_zero_df),
        "last_mile": build_last_mile_table(last_mile_df),
        "return_last_mile": build_last_mile_table(return_last_mile_df),
        "zone_map": zone_map,
    }


def get_simple_tariff(brackets: List[dict], liters: float) -> float:
    liters = max(liters, 0.0)
    for row in brackets:
        if value_in_bracket(liters, row["low"], row["high"]):
            if row["high"] is None and row["extra_per_l"] > 0:
                extra = max(liters - row["low"], 0.0)
                return row["base_tariff"] + math.ceil(extra) * row["extra_per_l"]
            return row["base_tariff"]
    return 0.0


def get_last_mile_tariff(table: List[dict], origin_zone: int, destination_label: str, liters: float) -> float:
    destination_label = clean_text(destination_label)
    liters = max(liters, 0.0)
    candidates = [r for r in table if r["origin_zone"] == origin_zone and r["dest_label"] == destination_label]
    for row in candidates:
        if value_in_bracket(liters, row["low"], row["high"]):
            if row["high"] is None and row["extra_per_l"] > 0:
                extra = max(liters - row["low"], 0.0)
                return row["base_tariff"] + math.ceil(extra) * row["extra_per_l"]
            return row["base_tariff"]
    return 0.0


def get_region_group_labels(logistics_dict: dict, origin_zone: int) -> List[str]:
    labels = []
    for row in logistics_dict["last_mile"]:
        if row["origin_zone"] == origin_zone:
            label = clean_text(row["dest_label"])
            if label and label not in labels:
                labels.append(label)
    return labels


def build_fbs_zone_scenarios(logistics_dict: dict, origin_zone: int, share_moscow: float, share_spb: float, share_regions: float):
    labels = get_region_group_labels(logistics_dict, origin_zone)
    regional_labels = [x for x in labels if x not in {"Москва и МО", "СПБ и ЛО"}]
    shares_total = max(share_moscow + share_spb + share_regions, 0.0001)
    wm = share_moscow / shares_total
    ws = share_spb / shares_total
    wr = share_regions / shares_total

    scenarios = {
        "Москва и МО": {"weights": {"Москва и МО": 1.0}, "kind": "single"},
        "СПБ и ЛО": {"weights": {"СПБ и ЛО": 1.0}, "kind": "single"},
    }

    if regional_labels:
        regional_weight = 1.0 / len(regional_labels)
        scenarios["Регионы (среднее)"] = {
            "weights": {label: regional_weight for label in regional_labels},
            "kind": "average_regions",
        }
    else:
        scenarios["Регионы (среднее)"] = {
            "weights": {},
            "kind": "average_regions",
        }

    weighted = {}
    if wm > 0:
        weighted["Москва и МО"] = wm
    if ws > 0:
        weighted["СПБ и ЛО"] = ws
    if wr > 0 and regional_labels:
        regional_weight = wr / len(regional_labels)
        for label in regional_labels:
            weighted[label] = weighted.get(label, 0.0) + regional_weight
    scenarios["Средневзвешенно"] = {"weights": weighted, "kind": "weighted"}
    return scenarios


# =========================
# Commission lookup
# =========================
def get_commission(row: pd.Series, commission_dict: Dict[str, Dict[str, float]]) -> Tuple[float, str]:
    manual = to_float(row.get("manual_commission"), -1)
    if manual >= 0:
        return (manual / 100.0 if manual > 1 else manual), "manual"

    template = clean_text(row.get("template")).lower()
    item_type = clean_text(row.get("item_type")).lower()
    subcategory = clean_text(row.get("subcategory")).lower()
    category = clean_text(row.get("category")).lower()

    if template and template in commission_dict["template"]:
        return commission_dict["template"][template], "template"
    if item_type and item_type in commission_dict["item_type"]:
        return commission_dict["item_type"][item_type], "item_type"
    if subcategory and subcategory in commission_dict["subcategory"]:
        return commission_dict["subcategory"][subcategory], "subcategory"
    if category and category in commission_dict["category"]:
        return commission_dict["category"][category], "category"
    return 0.0, "not_found"


# =========================
# Taxes
# =========================
def calc_tax(price: float, profit_before_tax: float, regime: str) -> Tuple[float, float, float]:
    price = max(price, 0.0)
    profit_before_tax = float(profit_before_tax)

    regimes = {
        "ОСНО (налог на прибыль 25%)": ("profit", 0.25),
        "УСН Доходы (6%)": ("revenue", 0.06),
        "УСН Доходы-Расходы (15%)": ("profit", 0.15),
        "АУСН Доходы (8%)": ("revenue", 0.08),
        "АУСН Доходы-Расходы (20%)": ("profit", 0.20),
        "Без налога": ("none", 0.0),
    }

    mode, rate = regimes.get(regime, ("none", 0.0))

    if mode == "revenue":
        tax = price * rate
    elif mode == "profit":
        tax = max(profit_before_tax, 0.0) * rate
    else:
        tax = 0.0

    profit_after_tax = profit_before_tax - tax
    margin_after_tax = (profit_after_tax / price * 100) if price > 0 else 0.0
    return round(tax, 2), round(profit_after_tax, 2), round(margin_after_tax, 2)


# =========================
# Core math
# =========================
def calc_expected_logistics_for_destination(
    scheme: str,
    chargeable_liters: float,
    origin_zone: int,
    destination_label: str,
    logistics_dict: dict,
    buyout_pct: float,
    client_return_pct: float,
    fbo_inbound_per_unit: float,
):
    zero_mile = get_simple_tariff(logistics_dict["zero_mile"], chargeable_liters)
    return_zero_mile = get_simple_tariff(logistics_dict["return_zero_mile"], chargeable_liters)
    last_mile = get_last_mile_tariff(logistics_dict["last_mile"], origin_zone, destination_label, chargeable_liters)
    return_last_mile = get_last_mile_tariff(logistics_dict["return_last_mile"], origin_zone, destination_label, chargeable_liters)

    buyout_rate = max(min(buyout_pct / 100.0, 0.9999), 0.0001)
    cancel_rate = 1.0 - buyout_rate
    client_return_rate = max(min(client_return_pct / 100.0, 1.0), 0.0)

    if scheme == "FBS":
        expected_zero_mile = zero_mile / buyout_rate
        expected_return_zero_mile = return_zero_mile * (cancel_rate / buyout_rate)
    else:
        expected_zero_mile = max(fbo_inbound_per_unit, 0.0)
        expected_return_zero_mile = 0.0

    expected_last_mile = last_mile / buyout_rate
    expected_return_last_mile = (
        return_last_mile * (cancel_rate / buyout_rate) +
        return_last_mile * client_return_rate
    )

    return {
        "expected_zero_mile": expected_zero_mile,
        "expected_return_zero_mile": expected_return_zero_mile,
        "expected_last_mile": expected_last_mile,
        "expected_return_last_mile": expected_return_last_mile,
        "logistics_total": expected_zero_mile + expected_return_zero_mile + expected_last_mile + expected_return_last_mile,
    }


def build_metrics_for_price(price: float, fixed_costs: float, variable_pct: float, cost_price: float, tax_regime: str):
    price = max(price, 0.0)
    variable_costs = price * variable_pct
    profit_before_tax = price - fixed_costs - variable_costs
    margin_before_tax = (profit_before_tax / price * 100) if price > 0 else 0.0
    tax, profit_after_tax, margin_after_tax = calc_tax(price, profit_before_tax, tax_regime)
    markup_pct = ((price / cost_price - 1) * 100) if cost_price > 0 else 0.0
    return {
        "price": round(price, 2),
        "variable_costs": round(variable_costs, 2),
        "profit_before_tax": round(profit_before_tax, 2),
        "margin_before_tax": round(margin_before_tax, 2),
        "tax": tax,
        "profit_after_tax": profit_after_tax,
        "margin_after_tax": margin_after_tax,
        "markup_pct": round(markup_pct, 2),
    }


def calculate_unit_metrics(
    row: pd.Series,
    scheme: str,
    tax_regime: str,
    commission_dict: Dict[str, Dict[str, float]],
    logistics_dict: dict,
    origin_zone: int,
    acquiring_pct: float,
    payout_pct: float,
    marketing_pct: float,
    other_mp_pct: float,
    packing_rub: float,
    other_fixed_rub: float,
    fbo_inbound_per_unit: float,
    buyout_pct: float,
    client_return_pct: float,
    target_margin_pct: float,
    round_price_step: int,
    zone_shares: Tuple[float, float, float],
) -> dict:
    commission_pct, commission_source = get_commission(row, commission_dict)

    cost_price = to_float(row.get("cost_price"), 0.0)
    current_price = to_float(row.get("current_price"), 0.0)
    promo_price = to_float(row.get("promo_price"), 0.0)

    length_cm = to_float(row.get("length_cm"), 0.0)
    width_cm = to_float(row.get("width_cm"), 0.0)
    height_cm = to_float(row.get("height_cm"), 0.0)
    weight_kg = to_float(row.get("weight_kg"), 0.0)

    liters = liters_from_dimensions(length_cm, width_cm, height_cm)
    chargeable_liters = liters if liters > 0 else weight_kg

    variable_pct = commission_pct + acquiring_pct / 100.0 + payout_pct / 100.0 + marketing_pct / 100.0 + other_mp_pct / 100.0
    base_price = promo_price if promo_price > 0 else current_price

    zone_data = {}

    if scheme == "FBS":
        scenario_map = build_fbs_zone_scenarios(
            logistics_dict,
            origin_zone,
            zone_shares[0],
            zone_shares[1],
            zone_shares[2],
        )

        for scenario_name, scenario in scenario_map.items():
            weights = scenario["weights"]
            if not weights:
                logistics_total = 0.0
                ez = erz = el = erl = 0.0
            else:
                weighted_parts = []
                for dest_label, weight in weights.items():
                    lg = calc_expected_logistics_for_destination(
                        scheme=scheme,
                        chargeable_liters=chargeable_liters,
                        origin_zone=origin_zone,
                        destination_label=dest_label,
                        logistics_dict=logistics_dict,
                        buyout_pct=buyout_pct,
                        client_return_pct=client_return_pct,
                        fbo_inbound_per_unit=fbo_inbound_per_unit,
                    )
                    weighted_parts.append((lg, weight))
                ez = sum(item[0]["expected_zero_mile"] * item[1] for item in weighted_parts)
                erz = sum(item[0]["expected_return_zero_mile"] * item[1] for item in weighted_parts)
                el = sum(item[0]["expected_last_mile"] * item[1] for item in weighted_parts)
                erl = sum(item[0]["expected_return_last_mile"] * item[1] for item in weighted_parts)
                logistics_total = sum(item[0]["logistics_total"] * item[1] for item in weighted_parts)

            fixed_costs = cost_price + logistics_total + packing_rub + other_fixed_rub
            denom = 1.0 - variable_pct - target_margin_pct / 100.0
            recommended_price_raw = fixed_costs / denom if denom > 0 else 0.0
            recommended_price = round_to_step(recommended_price_raw, round_price_step)

            current_metrics = build_metrics_for_price(base_price, fixed_costs, variable_pct, cost_price, tax_regime)
            recommended_metrics = build_metrics_for_price(recommended_price, fixed_costs, variable_pct, cost_price, tax_regime)
            zone_data[scenario_name] = {
                "zero_mile": round(ez, 2),
                "return_zero_mile": round(erz, 2),
                "last_mile": round(el, 2),
                "return_last_mile": round(erl, 2),
                "logistics_total": round(logistics_total, 2),
                "recommended_price": round(recommended_price, 2),
                "recommended_price_raw": round(recommended_price_raw, 2),
                "current_metrics": current_metrics,
                "recommended_metrics": recommended_metrics,
                "denom": denom,
            }

        primary = zone_data.get("Средневзвешенно", zone_data.get("Москва и МО"))
    else:
        lg = calc_expected_logistics_for_destination(
            scheme=scheme,
            chargeable_liters=chargeable_liters,
            origin_zone=origin_zone,
            destination_label="Москва и МО",
            logistics_dict=logistics_dict,
            buyout_pct=buyout_pct,
            client_return_pct=client_return_pct,
            fbo_inbound_per_unit=fbo_inbound_per_unit,
        )
        logistics_total = lg["logistics_total"]
        fixed_costs = cost_price + logistics_total + packing_rub + other_fixed_rub
        denom = 1.0 - variable_pct - target_margin_pct / 100.0
        recommended_price_raw = fixed_costs / denom if denom > 0 else 0.0
        recommended_price = round_to_step(recommended_price_raw, round_price_step)

        primary = {
            "zero_mile": round(lg["expected_zero_mile"], 2),
            "return_zero_mile": round(lg["expected_return_zero_mile"], 2),
            "last_mile": round(lg["expected_last_mile"], 2),
            "return_last_mile": round(lg["expected_return_last_mile"], 2),
            "logistics_total": round(logistics_total, 2),
            "recommended_price": round(recommended_price, 2),
            "recommended_price_raw": round(recommended_price_raw, 2),
            "current_metrics": build_metrics_for_price(base_price, fixed_costs, variable_pct, cost_price, tax_regime),
            "recommended_metrics": build_metrics_for_price(recommended_price, fixed_costs, variable_pct, cost_price, tax_regime),
            "denom": denom,
        }
        zone_data = {
            "Средневзвешенно": primary,
            "Москва и МО": primary,
            "СПБ и ЛО": primary,
            "Регионы (среднее)": primary,
        }

    bad_flag = ""
    if commission_source == "not_found":
        bad_flag = "Нет комиссии"
    elif primary["denom"] <= 0:
        bad_flag = "Целевая маржа недостижима"
    elif primary["current_metrics"]["profit_after_tax"] < 0:
        bad_flag = "Убыточно"
    elif primary["current_metrics"]["margin_before_tax"] < target_margin_pct:
        bad_flag = "Ниже цели"

    weighted_current = primary["current_metrics"]
    weighted_recommended = primary["recommended_metrics"]

    return {
        "SKU": clean_text(row.get("sku")),
        "Название": clean_text(row.get("name")),
        "Схема": scheme,
        "Объем, л": round(chargeable_liters, 3),
        "Вес, кг": round(weight_kg, 3),
        "Себестоимость, руб": round(cost_price, 2),
        "Текущая цена, руб": round(current_price, 2),
        "Цена акции, руб": round(promo_price, 2),
        "Цена расчета, руб": weighted_current["price"],
        "Комиссия, %": round(commission_pct * 100, 2),
        "Источник комиссии": commission_source,
        "Нулевая миля (средняя), руб": primary["zero_mile"],
        "Возврат нулевой мили (средняя), руб": primary["return_zero_mile"],
        "Последняя миля (средняя), руб": primary["last_mile"],
        "Возврат последней мили (средняя), руб": primary["return_last_mile"],
        "Логистика итого (средняя), руб": primary["logistics_total"],
        "Переменные расходы, руб": weighted_current["variable_costs"],
        "Прибыль до налога (текущая), руб": weighted_current["profit_before_tax"],
        "Маржа до налога (текущая), %": weighted_current["margin_before_tax"],
        "Налог (текущая), руб": weighted_current["tax"],
        "Прибыль после налога (текущая), руб": weighted_current["profit_after_tax"],
        "Маржа после налога (текущая), %": weighted_current["margin_after_tax"],
        "Наценка (текущая), %": weighted_current["markup_pct"],
        "Рекоменд. цена Москва и МО, руб": zone_data["Москва и МО"]["recommended_price"],
        "Рекоменд. цена СПБ и ЛО, руб": zone_data["СПБ и ЛО"]["recommended_price"],
        "Рекоменд. цена Регионы, руб": zone_data["Регионы (среднее)"]["recommended_price"],
        "Рекоменд. цена средняя, руб": zone_data["Средневзвешенно"]["recommended_price"],
        "Логистика Москва и МО, руб": zone_data["Москва и МО"]["logistics_total"],
        "Логистика СПБ и ЛО, руб": zone_data["СПБ и ЛО"]["logistics_total"],
        "Логистика Регионы, руб": zone_data["Регионы (среднее)"]["logistics_total"],
        "Рекоменд. цена к публикации, руб": weighted_recommended["price"],
        "Прибыль до налога (рекоменд.), руб": weighted_recommended["profit_before_tax"],
        "Маржа до налога (рекоменд.), %": weighted_recommended["margin_before_tax"],
        "Налог (рекоменд.), руб": weighted_recommended["tax"],
        "Прибыль после налога (рекоменд.), руб": weighted_recommended["profit_after_tax"],
        "Маржа после налога (рекоменд.), %": weighted_recommended["margin_after_tax"],
        "Наценка (рекоменд.), %": weighted_recommended["markup_pct"],
        "Флаг": bad_flag,
    }


# =========================
# Excel export
# =========================
def autofit_worksheet(ws):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            widths[cell.column] = max(widths.get(cell.column, 0), min(len(value) + 2, 40))
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def apply_table_style(ws, freeze_cell="A2"):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    ws.freeze_panes = freeze_cell
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def set_number_formats(ws):
    currency_keywords = ["руб", "цена", "логистика", "прибыль", "налог", "себестоимость", "расходы"]
    pct_keywords = ["%", "маржа", "комиссия", "наценка"]
    for col_idx, cell in enumerate(ws[1], start=1):
        header = str(cell.value).lower() if cell.value else ""
        if any(k in header for k in currency_keywords):
            fmt = '#,##0.00;[Red](#,##0.00);-'
        elif any(k in header for k in pct_keywords):
            fmt = '0.00;[Red](0.00);-'
        else:
            fmt = None
        if fmt:
            for data_cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, max_row=ws.max_row):
                for c in data_cell:
                    if isinstance(c.value, (int, float)):
                        c.number_format = fmt


def write_dataframe_sheet(ws, df: pd.DataFrame, sheet_name: str):
    ws.title = sheet_name
    ws.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))
    apply_table_style(ws)
    set_number_formats(ws)
    autofit_worksheet(ws)


def dataframe_to_excel_bytes(result_df: pd.DataFrame, template_df: pd.DataFrame, kpi_df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws_kpi = wb.active
    write_dataframe_sheet(ws_kpi, kpi_df, "KPI")
    ws_results = wb.create_sheet("Результат")
    write_dataframe_sheet(ws_results, result_df, "Результат")
    ws_template = wb.create_sheet("Шаблон")
    write_dataframe_sheet(ws_template, template_df, "Шаблон")

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# =========================
# UI
# =========================
conn = init_db()

st.title("Лемана Про — единый калькулятор юнит-экономики")
st.caption("Поддержка схем FBS и FBO. Комиссии и тарифы загружаются из актуальных Excel-файлов Лемана Про.")

with st.sidebar:
    st.header("Параметры модели")

    scheme = st.radio("Схема работы", ["FBS", "FBO"], horizontal=True)

    tax_regime = st.selectbox(
        "Налоговый режим",
        [
            "ОСНО (налог на прибыль 25%)",
            "УСН Доходы (6%)",
            "УСН Доходы-Расходы (15%)",
            "АУСН Доходы (8%)",
            "АУСН Доходы-Расходы (20%)",
            "Без налога",
        ],
    )

    st.divider()
    st.subheader("Процентные расходы")
    target_margin_pct = st.slider("Целевая маржа до налога, %", 0, 80, 20)
    acquiring_pct = st.number_input("Эквайринг, %", min_value=0.0, max_value=10.0, value=1.5, step=0.1)
    payout_pct = st.number_input("Ранняя выплата / фин. услуги, %", min_value=0.0, max_value=10.0, value=0.0, step=0.1)
    marketing_pct = st.number_input("Маркетинг / реклама, %", min_value=0.0, max_value=50.0, value=5.0, step=0.1)
    other_mp_pct = st.number_input("Прочие % маркетплейса, %", min_value=0.0, max_value=50.0, value=0.0, step=0.1)

    st.divider()
    st.subheader("Фиксированные расходы")
    packing_rub = st.number_input("Упаковка на ед., руб", min_value=0.0, max_value=5000.0, value=0.0, step=10.0)
    other_fixed_rub = st.number_input("Прочие фикс. расходы на ед., руб", min_value=0.0, max_value=5000.0, value=0.0, step=10.0)
    if scheme == "FBO":
        fbo_inbound_per_unit = st.number_input(
            "Входящая логистика FBO на ед., руб",
            min_value=0.0,
            max_value=10000.0,
            value=0.0,
            step=10.0,
            help="Для FBO обычно корректнее задавать усредненную входящую логистику на единицу отдельно.",
        )
    else:
        fbo_inbound_per_unit = 0.0

    st.divider()
    st.subheader("Поведенческие параметры")
    buyout_pct = st.slider("Выкуп, %", 1, 100, 95)
    client_return_pct = st.slider("Возвраты после выкупа, %", 0, 100, 5)

    st.divider()
    st.subheader("Логистика")
    origin_zone = st.number_input("Зона откуда", min_value=1, max_value=10, value=9, step=1, help="Для склада Москва / МО обычно это зона 9, если тарифный файл не менялся.")
    round_price_step = st.selectbox("Округлять рекомендованную цену до", [1, 5, 10, 50, 100], index=2)

    if scheme == "FBS":
        st.divider()
        st.subheader("Средняя модель зон FBS")
        share_moscow = st.number_input("Доля заказов Москва и МО, %", min_value=0.0, max_value=100.0, value=70.0, step=1.0)
        share_spb = st.number_input("Доля заказов СПБ и ЛО, %", min_value=0.0, max_value=100.0, value=20.0, step=1.0)
        share_regions = st.number_input("Доля заказов регионы, %", min_value=0.0, max_value=100.0, value=10.0, step=1.0)
    else:
        share_moscow, share_spb, share_regions = 0.0, 0.0, 0.0

st.markdown("### 1. Файлы тарифов и комиссий")
col_a, col_b = st.columns(2)
with col_a:
    commission_file = st.file_uploader(
        "Файл комиссий Лемана Про",
        type=["xlsx"],
        key="commission_file",
        help="Нужен лист 'Комиссия_FBS и FBO'.",
    )
with col_b:
    logistics_file = st.file_uploader(
        "Файл логистики Лемана Про",
        type=["xlsx"],
        key="logistics_file",
        help="Нужны листы 'Тарифы (Последняя миля)', 'Тарифы (возврат Последняя миля)', 'Тарифы (Доставка до СЦ)', 'Тарифы (Возврат Доставка до СЦ)', 'Зоны (услуга Последняя миля)'.",
    )

if not commission_file or not logistics_file:
    st.info("Сначала загрузите оба файла Лемана Про: комиссии и логистику.")
    st.stop()

try:
    commission_dict = load_commissions(commission_file)
    logistics_dict = load_logistics(logistics_file)
except Exception as e:
    st.error(f"Не удалось прочитать тарифные файлы: {e}")
    st.stop()

st.success("Файлы Лемана Про успешно загружены.")

st.markdown("### 2. Каталог товаров")
dim_col, wt_col = st.columns(2)
with dim_col:
    dim_unit = st.selectbox("Единица размеров в файле каталога", ["см", "мм", "м"], index=0)
with wt_col:
    wt_unit = st.selectbox("Единица веса в файле каталога", ["кг", "г"], index=0)

catalog_file = st.file_uploader(
    "Загрузить каталог Excel",
    type=["xlsx"],
    key="catalog_file",
    help="Поддерживаются: SKU/Артикул, Название/Наименование, Шаблон товара, Тип товара, Подкатегория товара, Категория, Длина, Ширина, Высота, Вес, Себестоимость, Текущая цена, Цена акции, Регион.",
)

if catalog_file:
    try:
        df = pd.read_excel(catalog_file)
        save_rows = []
        for _, row in df.iterrows():
            sku = clean_text(first_existing(row, ["SKU", "Артикул", "Артикул продавца"]))
            name = clean_text(first_existing(row, ["Название", "Наименование", "Товар"]))
            if not sku:
                continue

            save_rows.append(
                (
                    sku,
                    name,
                    clean_text(first_existing(row, ["Шаблон товара", "Шаблон"])),
                    clean_text(first_existing(row, ["Тип товара"])),
                    clean_text(first_existing(row, ["Подкатегория товара", "Подкатегория"])),
                    clean_text(first_existing(row, ["Категория"])),
                    normalize_dimension(first_existing(row, ["Длина", "Длина, см", "Длина, мм"]), dim_unit),
                    normalize_dimension(first_existing(row, ["Ширина", "Ширина, см", "Ширина, мм"]), dim_unit),
                    normalize_dimension(first_existing(row, ["Высота", "Высота, см", "Высота, мм"]), dim_unit),
                    normalize_weight(first_existing(row, ["Вес", "Вес, кг", "Вес, г"]), wt_unit),
                    to_float(first_existing(row, ["Себестоимость", "Себес", "Закупка"]), 0.0),
                    to_float(first_existing(row, ["Текущая цена", "Цена", "Цена без акции"]), 0.0),
                    to_float(first_existing(row, ["Цена акции", "Акционная цена", "Цена со скидкой"]), 0.0),
                    clean_text(first_existing(row, ["Регион", "Область", "Край/Область"])),
                    None,
                )
            )

        if save_rows:
            conn.executemany(
                """
                INSERT INTO products (
                    sku, name, template, item_type, subcategory, category,
                    length_cm, width_cm, height_cm, weight_kg,
                    cost_price, current_price, promo_price, region, manual_commission
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(sku) DO UPDATE SET
                    name=excluded.name,
                    template=excluded.template,
                    item_type=excluded.item_type,
                    subcategory=excluded.subcategory,
                    category=excluded.category,
                    length_cm=excluded.length_cm,
                    width_cm=excluded.width_cm,
                    height_cm=excluded.height_cm,
                    weight_kg=excluded.weight_kg,
                    cost_price=excluded.cost_price,
                    current_price=excluded.current_price,
                    promo_price=excluded.promo_price,
                    region=excluded.region
                """,
                save_rows,
            )
            conn.commit()
            st.success(f"Загружено / обновлено SKU: {len(save_rows)}")
    except Exception as e:
        st.error(f"Ошибка загрузки каталога: {e}")

catalog_df = pd.read_sql_query("SELECT * FROM products ORDER BY sku", conn)

if catalog_df.empty:
    st.warning("Каталог пока пуст. Загрузите Excel с товарами.")
    st.stop()

st.markdown("#### Ручная корректировка каталога и комиссий")
editable = catalog_df.copy()
editable["manual_commission"] = editable["manual_commission"].fillna("")
edited = st.data_editor(
    editable,
    use_container_width=True,
    num_rows="dynamic",
    hide_index=True,
    column_config={
        "manual_commission": st.column_config.NumberColumn("Комиссия вручную, %", min_value=0.0, max_value=100.0, step=0.1),
        "current_price": st.column_config.NumberColumn("Текущая цена, руб", min_value=0.0, step=1.0),
        "promo_price": st.column_config.NumberColumn("Цена акции, руб", min_value=0.0, step=1.0),
        "cost_price": st.column_config.NumberColumn("Себестоимость, руб", min_value=0.0, step=1.0),
        "length_cm": st.column_config.NumberColumn("Длина, см", min_value=0.0, step=0.1),
        "width_cm": st.column_config.NumberColumn("Ширина, см", min_value=0.0, step=0.1),
        "height_cm": st.column_config.NumberColumn("Высота, см", min_value=0.0, step=0.1),
        "weight_kg": st.column_config.NumberColumn("Вес, кг", min_value=0.0, step=0.001),
    },
)

if st.button("Сохранить изменения в каталог", type="secondary"):
    rows = []
    for _, row in edited.iterrows():
        rows.append(
            (
                clean_text(row["name"]),
                clean_text(row["template"]),
                clean_text(row["item_type"]),
                clean_text(row["subcategory"]),
                clean_text(row["category"]),
                to_float(row["length_cm"], 0.0),
                to_float(row["width_cm"], 0.0),
                to_float(row["height_cm"], 0.0),
                to_float(row["weight_kg"], 0.0),
                to_float(row["cost_price"], 0.0),
                to_float(row["current_price"], 0.0),
                to_float(row["promo_price"], 0.0),
                clean_text(row["region"]),
                None if clean_text(row["manual_commission"]) == "" else to_float(row["manual_commission"], 0.0),
                clean_text(row["sku"]),
            )
        )
    conn.executemany(
        """
        UPDATE products
        SET name=?, template=?, item_type=?, subcategory=?, category=?,
            length_cm=?, width_cm=?, height_cm=?, weight_kg=?,
            cost_price=?, current_price=?, promo_price=?, region=?, manual_commission=?
        WHERE sku=?
        """,
        rows,
    )
    conn.commit()
    st.success("Изменения сохранены.")
    catalog_df = pd.read_sql_query("SELECT * FROM products ORDER BY sku", conn)

st.markdown("### 3. Расчёт")

if st.button("Рассчитать юнит-экономику", type="primary"):
    current_catalog = pd.read_sql_query("SELECT * FROM products ORDER BY sku", conn)

    results = []
    for _, row in current_catalog.iterrows():
        results.append(
            calculate_unit_metrics(
                row=row,
                scheme=scheme,
                tax_regime=tax_regime,
                commission_dict=commission_dict,
                logistics_dict=logistics_dict,
                origin_zone=int(origin_zone),
                acquiring_pct=acquiring_pct,
                payout_pct=payout_pct,
                marketing_pct=marketing_pct,
                other_mp_pct=other_mp_pct,
                packing_rub=packing_rub,
                other_fixed_rub=other_fixed_rub,
                fbo_inbound_per_unit=fbo_inbound_per_unit,
                buyout_pct=buyout_pct,
                client_return_pct=client_return_pct,
                target_margin_pct=target_margin_pct,
                round_price_step=int(round_price_step),
                zone_shares=(share_moscow, share_spb, share_regions),
            )
        )

    result_df = pd.DataFrame(results)

    total_sku = len(result_df)
    profitable_count = int((result_df["Прибыль после налога (текущая), руб"] > 0).sum())
    loss_count = int((result_df["Прибыль после налога (текущая), руб"] <= 0).sum())
    avg_margin = round(result_df["Маржа после налога (текущая), %"].mean(), 2) if total_sku else 0.0
    avg_recommended_price = round(result_df["Рекоменд. цена к публикации, руб"].mean(), 2) if total_sku else 0.0
    avg_rec_moscow = round(result_df["Рекоменд. цена Москва и МО, руб"].mean(), 2) if total_sku else 0.0
    avg_rec_spb = round(result_df["Рекоменд. цена СПБ и ЛО, руб"].mean(), 2) if total_sku else 0.0
    avg_rec_regions = round(result_df["Рекоменд. цена Регионы, руб"].mean(), 2) if total_sku else 0.0

    k1, k2, k3, k4, k5 = st.columns(5)
    k1.metric("SKU в расчёте", total_sku)
    k2.metric("Плюсовых SKU", profitable_count)
    k3.metric("Убыточных / нулевых SKU", loss_count)
    k4.metric("Средняя маржа после налога, %", avg_margin)
    k5.metric("Средняя рекоменд. цена", avg_recommended_price)

    st.markdown("#### Средние рекомендованные цены по зонам")
    z1, z2, z3, z4 = st.columns(4)
    z1.metric("Москва и МО", avg_rec_moscow)
    z2.metric("СПБ и ЛО", avg_rec_spb)
    z3.metric("Регионы", avg_rec_regions)
    z4.metric("Одна цена к публикации", avg_recommended_price)

    st.caption("Для FBS цена к публикации считается как средневзвешенная по долям зон. Для FBO значения по зонам совпадают, потому что используется единая модель входящей логистики на единицу.")

    def color_flags(val):
        if val in ("Убыточно", "Нет комиссии", "Целевая маржа недостижима"):
            return "background-color: #ffdddd"
        if val == "Ниже цели":
            return "background-color: #fff2cc"
        return ""

    styled = result_df.style.applymap(color_flags, subset=["Флаг"])
    st.dataframe(styled, use_container_width=True)

    template_cols = [
        "SKU",
        "Название",
        "Себестоимость, руб",
        "Текущая цена, руб",
        "Цена акции, руб",
        "Комиссия, %",
        "Источник комиссии",
        "Логистика Москва и МО, руб",
        "Логистика СПБ и ЛО, руб",
        "Логистика Регионы, руб",
        "Логистика итого (средняя), руб",
        "Рекоменд. цена Москва и МО, руб",
        "Рекоменд. цена СПБ и ЛО, руб",
        "Рекоменд. цена Регионы, руб",
        "Рекоменд. цена средняя, руб",
        "Рекоменд. цена к публикации, руб",
        "Прибыль после налога (текущая), руб",
        "Маржа после налога (текущая), %",
        "Флаг",
    ]
    notes_df = result_df[template_cols].copy()
    notes_df["Комментарий для сотрудника"] = notes_df["Флаг"].map(
        {
            "Убыточно": "Проверь цену, комиссию и логистику: SKU сейчас убыточен.",
            "Ниже цели": "SKU прибыльный, но не дотягивает до целевой маржи.",
            "Нет комиссии": "Нужно вручную проставить категорию или комиссию.",
            "Целевая маржа недостижима": "Переменные расходы слишком высокие для выбранной целевой маржи.",
        }
    ).fillna("SKU в норме.")

    kpi_df = pd.DataFrame(
        [
            {"Показатель": "SKU в расчёте", "Значение": total_sku},
            {"Показатель": "Плюсовых SKU", "Значение": profitable_count},
            {"Показатель": "Убыточных / нулевых SKU", "Значение": loss_count},
            {"Показатель": "Средняя маржа после налога, %", "Значение": avg_margin},
            {"Показатель": "Средняя рекоменд. цена Москва и МО, руб", "Значение": avg_rec_moscow},
            {"Показатель": "Средняя рекоменд. цена СПБ и ЛО, руб", "Значение": avg_rec_spb},
            {"Показатель": "Средняя рекоменд. цена Регионы, руб", "Значение": avg_rec_regions},
            {"Показатель": "Средняя цена к публикации, руб", "Значение": avg_recommended_price},
            {"Показатель": "Схема", "Значение": scheme},
            {"Показатель": "Зона откуда", "Значение": int(origin_zone)},
        ]
    )

    result_xlsx = dataframe_to_excel_bytes(result_df, notes_df, kpi_df)
    st.download_button(
        "Скачать результат Excel (.xlsx)",
        data=result_xlsx,
        file_name=f"lemanpro_unit_economics_{scheme.lower()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    template_xlsx = dataframe_to_excel_bytes(notes_df, notes_df, kpi_df)
    st.download_button(
        "Скачать шаблон для сотрудников Excel (.xlsx)",
        data=template_xlsx,
        file_name=f"lemanpro_template_for_team_{scheme.lower()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
else:
    st.info("После загрузки файлов и каталога нажмите «Рассчитать юнит-экономику».")
